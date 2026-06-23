# -*- coding: utf-8 -*-
"""
Le a planilha modelo do SMT/SAC (aba "INSIRA OS DADOS") e devolve o conteudo
no MESMO formato de string que o fluxo .txt usa (campos separados por '#').

Por que ler a aba "INSIRA OS DADOS" e nao a "TRANSFORME EM TXT":
    A aba "TRANSFORME EM TXT" depende de formulas do Excel para concatenar os
    campos. Se a planilha for reaberta/salva em outro programa (LibreOffice),
    ou uma formula quebrar, a string sai corrompida sem aviso. Lendo as colunas
    cruas da aba de dados, montamos a string '#' aqui em Python, com controle
    total e sem depender de formula nenhuma.

Mapeamento (confirmado e VALIDADO contra a planilha modelo):
    A string '#' NAO e a juncao simples das 27 colunas. A formula original
    insere UM CAMPO VAZIO FIXO entre o bloco de cadastro (apelido..fax, 18
    campos) e o bloco de auxiliares (9). O resultado tem 28 campos + '#' final:

        idx0  apelido        idx9   cep            idx18 (VAZIO FIXO)
        idx1  titulo         idx10  pais           idx19 auxiliar1
        idx2  nome           idx11  provincia      idx20 auxiliar2
        idx3  endereco       idx12  cep_cxpostal   idx21 auxiliar3
        idx4  bairro         idx13  ddd            idx22 auxiliar4
        idx5  complemento    idx14  telefone       idx23 auxiliar5
        idx6  numero         idx15  email          idx24 auxiliar6
        idx7  cidade         idx16  ddd_fax        idx25 auxiliar7
        idx8  uf             idx17  fax            idx26 auxiliar8
                                                   idx27 auxiliar9

Esse campo vazio no idx18 e CRITICO: txt_para_correios.linha_txt_para_campos()
le auxiliar1 no indice 19, auxiliar2 no 20, ... Sem o campo vazio, todos os
auxiliares desalinham e o CSV sai com auxiliares no campo errado, SEM ERRO
visivel. Por isso montamos: 18 colunas + '' + 9 colunas de auxiliar.
"""

ABA_DADOS = 'INSIRA OS DADOS'
# Layout da planilha INSIRA: 18 colunas de cadastro + 9 de auxiliar = 27.
N_CADASTRO = 18         # apelido .. fax  (colunas 0..17)
N_AUXILIAR = 9          # auxiliar1 .. auxiliar9  (colunas 18..26)
N_COLUNAS = N_CADASTRO + N_AUXILIAR   # 27 colunas na aba
COL_NOME = 2            # se vazia, a linha e ignorada (igual ao fluxo .txt)


def _valor_celula(cell, datemode):
    """Converte uma celula em string limpa, tratando os tipos do .xls:
    numeros inteiros (CEP, numero, DDD) nao podem virar '18800053.0'."""
    import xlrd
    t = cell.ctype
    v = cell.value
    if t == xlrd.XL_CELL_EMPTY or t == xlrd.XL_CELL_BLANK:
        return ''
    if t == xlrd.XL_CELL_NUMBER:
        # inteiro -> sem '.0'; decimal real -> mantem
        if float(v).is_integer():
            return str(int(v))
        return str(v)
    if t == xlrd.XL_CELL_DATE:
        # data -> dd/mm/aaaa (raro nesta planilha, mas seguro)
        y, m, d, *_ = xlrd.xldate_as_tuple(v, datemode)
        return '{:02d}/{:02d}/{:04d}'.format(d, m, y)
    if t == xlrd.XL_CELL_BOOLEAN:
        return '1' if v else '0'
    return str(v).strip()


def _linhas_de_sheet(sheet, datemode):
    """Gera as strings '#' a partir das linhas de dados de uma aba .xls/.xlsx,
    pulando o cabecalho (linha 0) e linhas sem nome.

    Monta: 18 campos de cadastro + 1 campo VAZIO FIXO + 9 auxiliares + '#'.
    O campo vazio e o que a formula original insere e o que o
    txt_para_correios espera (auxiliar1 no indice 19)."""
    linhas = []
    for r in range(1, sheet.nrows):
        # le ate N_COLUNAS, completando com '' se a linha tiver menos colunas
        campos = []
        for c in range(N_COLUNAS):
            if c < sheet.ncols:
                campos.append(_valor_celula(sheet.cell(r, c), datemode))
            else:
                campos.append('')
        if not campos[COL_NOME].strip():
            continue  # linha vazia / sem nome -> ignora
        cadastro = campos[:N_CADASTRO]      # apelido .. fax
        auxiliar = campos[N_CADASTRO:]      # auxiliar1 .. auxiliar9
        # cadastro + campo vazio fixo + auxiliares, terminando em '#'
        linhas.append('#'.join(cadastro + [''] + auxiliar) + '#')
    return linhas


def ler_planilha_bytes(dados, nome_arquivo=''):
    """Recebe os BYTES de um .xls/.xlsx e devolve uma string multi-linha
    no formato .txt (cada registro numa linha, campos separados por '#').

    Le SOMENTE a aba 'INSIRA OS DADOS'. Se ela nao existir, levanta erro
    claro em vez de adivinhar outra aba."""
    nome = (nome_arquivo or '').lower()

    if nome.endswith('.xlsx'):
        return _ler_xlsx(dados)
    # default: .xls (formato antigo, usado pela planilha modelo atual)
    return _ler_xls(dados)


def _ler_xls(dados):
    import xlrd
    wb = xlrd.open_workbook(file_contents=dados)
    if ABA_DADOS not in wb.sheet_names():
        raise ValueError(
            "A planilha nao tem a aba '{}'. Abas encontradas: {}".format(
                ABA_DADOS, ', '.join(wb.sheet_names())))
    sheet = wb.sheet_by_name(ABA_DADOS)
    linhas = _linhas_de_sheet(sheet, wb.datemode)
    return '\n'.join(linhas)


def _ler_xlsx(dados):
    """Suporte a .xlsx via openpyxl (caso a planilha seja salva no formato novo)."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(dados), read_only=True, data_only=True)
    if ABA_DADOS not in wb.sheetnames:
        raise ValueError(
            "A planilha nao tem a aba '{}'. Abas encontradas: {}".format(
                ABA_DADOS, ', '.join(wb.sheetnames)))
    ws = wb[ABA_DADOS]
    linhas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # cabecalho
        campos = []
        for c in range(N_COLUNAS):
            v = row[c] if c < len(row) and row[c] is not None else ''
            if isinstance(v, float) and v.is_integer():
                v = str(int(v))
            campos.append(str(v).strip())
        if not campos[COL_NOME].strip():
            continue
        cadastro = campos[:N_CADASTRO]
        auxiliar = campos[N_CADASTRO:]
        linhas.append('#'.join(cadastro + [''] + auxiliar) + '#')
    return '\n'.join(linhas)
